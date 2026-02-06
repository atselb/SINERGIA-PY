from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional

from openpyxl import Workbook, load_workbook

@dataclass(frozen=True)
class ExcelConfig:
    path: Path
    sheet: str
    header_row: int = 1
    start_row: int = 2
    stop_when_first_cell_empty: bool = True

def _headers(ws, header_row: int) -> List[str]:
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in row]
    if any(h == "" for h in headers):
        raise ValueError("Headers inválidos: hay celdas vacías en la fila de headers.")
    return headers

def iter_rows(cfg: ExcelConfig) -> Iterator[Dict[str, Any]]:
    wb = load_workbook(cfg.path, read_only=True, data_only=True)
    if cfg.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{cfg.sheet}' no existe. Disponibles: {wb.sheetnames}")
    ws = wb[cfg.sheet]

    headers = _headers(ws, cfg.header_row)

    for row in ws.iter_rows(min_row=cfg.start_row, values_only=True):
        first = row[0] if len(row) else None
        if cfg.stop_when_first_cell_empty and (first is None or str(first).strip() == ""):
            break
        yield {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}

def write_rows_xlsx(
    path: Path,
    sheet: str,
    rows: Iterable[Dict[str,Any]],
    headers: Optional[List[str]] = None,
    overwrite: bool = True,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists() and not overwrite:
        raise FileExistsError(f"El archivo '{path}' ya existe y overwrite es False.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    rows_list = list(rows)

    if headers is None:

        headers = list(rows_list[0].keys()) if rows_list else []

    ws.append(headers)

    for r in rows_list:
        ws.append([r.get(h) for h in headers])
    
    wb.save(path)
